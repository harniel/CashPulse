"""
CSV import pipeline (Section 16): upload+validate happen together in
create_batch() (no pandas — stdlib csv is plenty at this scale), then
confirm_batch() persists whichever rows the user approved. Duplicate
detection flags, never auto-skips — the user decides per row.
"""

import csv
import io
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

import openpyxl
from django.db import transaction as db_transaction
from rest_framework.exceptions import PermissionDenied, ValidationError

from budgets import services as budget_services
from budgets.models import Budget
from categories.models import Category
from households.models import Household
from transactions.models import Transaction

from .models import BudgetImportBatch, BudgetImportRow, ImportBatch, ImportRow

MAX_ROWS = 500
MAX_FILE_SIZE_BYTES = 2 * 1024 * 1024  # 2MB
DUPLICATE_WINDOW_DAYS = 2
DATE_FORMATS = ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"]


def _parse_date(raw_value):
    raw_value = (raw_value or "").strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(raw_value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(raw_value):
    raw_value = (raw_value or "").strip().replace(",", "")
    try:
        amount = Decimal(raw_value)
    except InvalidOperation:
        return None
    if amount == 0:
        return None
    return amount


def _default_category(kind):
    # Bank CSV exports don't categorize — every imported row lands in the
    # matching system catch-all category (seeded by
    # categories/migrations/0002_seed_default_categories.py) and the user
    # recategorizes afterward through the normal Transaction PATCH endpoint,
    # same as any other transaction.
    name = "Other Income" if kind == Transaction.Type.INCOME else "Other Expense"
    return Category.objects.filter(is_system=True, kind=kind, name=name).first()


def _is_duplicate(account, row_date, amount):
    window_start = row_date - timedelta(days=DUPLICATE_WINDOW_DAYS)
    window_end = row_date + timedelta(days=DUPLICATE_WINDOW_DAYS)
    return Transaction.objects.filter(
        account=account, date__gte=window_start, date__lte=window_end, amount=amount
    ).exists()


def create_batch(user, account, uploaded_file, date_column, description_column, amount_column):
    if account.user_id != user.id:
        raise ValidationError({"account": "You don't have access to that account."})
    if not uploaded_file.name.lower().endswith(".csv"):
        raise ValidationError({"file": "Only .csv files are accepted."})
    if uploaded_file.size > MAX_FILE_SIZE_BYTES:
        raise ValidationError({"file": f"File exceeds the {MAX_FILE_SIZE_BYTES // 1024}KB limit."})

    try:
        decoded = uploaded_file.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        raise ValidationError({"file": "Could not read the file as UTF-8 text."})

    reader = csv.DictReader(io.StringIO(decoded))
    if not reader.fieldnames:
        raise ValidationError({"file": "The file has no header row."})
    for column in (date_column, description_column, amount_column):
        if column not in reader.fieldnames:
            raise ValidationError({"file": f"Column '{column}' not found in the file header."})

    rows = list(reader)
    if not rows:
        raise ValidationError({"file": "The file has no data rows."})
    if len(rows) > MAX_ROWS:
        raise ValidationError(
            {"file": f"This file has {len(rows)} rows; the limit for a single import is {MAX_ROWS}."}
        )

    with db_transaction.atomic():
        batch = ImportBatch.objects.create(
            user=user,
            account=account,
            filename=uploaded_file.name,
            row_count=len(rows),
            date_column=date_column,
            description_column=description_column,
            amount_column=amount_column,
        )
        for raw_row in rows:
            _stage_row(batch, account, raw_row, date_column, amount_column)

    return batch


def _stage_row(batch, account, raw_row, date_column, amount_column):
    row_date = _parse_date(raw_row.get(date_column))
    amount = _parse_amount(raw_row.get(amount_column))

    if row_date is None:
        ImportRow.objects.create(
            batch=batch,
            raw_data=raw_row,
            status=ImportRow.Status.FAILED,
            error=f"Couldn't parse '{raw_row.get(date_column)}' as a date.",
        )
        return
    if amount is None:
        ImportRow.objects.create(
            batch=batch,
            raw_data=raw_row,
            status=ImportRow.Status.FAILED,
            error=f"Couldn't parse '{raw_row.get(amount_column)}' as a nonzero amount.",
        )
        return

    ImportRow.objects.create(
        batch=batch,
        raw_data=raw_row,
        status=ImportRow.Status.PENDING,
        is_duplicate=_is_duplicate(account, row_date, abs(amount)),
    )


def confirm_batch(batch, row_ids=None):
    """
    Persists a Transaction for each selected pending row. `row_ids=None`
    means "everything not flagged as a duplicate" — an explicit list lets
    the caller include specific duplicates or exclude specific rows.
    Whatever's left pending afterward is marked SKIPPED; the batch can
    only be confirmed once.
    """
    if batch.status == ImportBatch.Status.CONFIRMED:
        raise ValidationError("This import has already been confirmed.")

    pending_rows = list(batch.rows.filter(status=ImportRow.Status.PENDING))
    pending_ids = {row.id for row in pending_rows}

    if row_ids is None:
        selected_ids = {row.id for row in pending_rows if not row.is_duplicate}
    else:
        selected_ids = set(row_ids)
        unknown = selected_ids - pending_ids
        if unknown:
            raise ValidationError(
                {"row_ids": f"Unknown or non-pending row id(s): {sorted(str(i) for i in unknown)}"}
            )

    imported = []
    with db_transaction.atomic():
        for row in pending_rows:
            if row.id not in selected_ids:
                row.status = ImportRow.Status.SKIPPED
                row.save(update_fields=["status"])
                continue

            row_date = _parse_date(row.raw_data.get(batch.date_column))
            amount = _parse_amount(row.raw_data.get(batch.amount_column))
            description = (row.raw_data.get(batch.description_column) or "").strip()
            type_ = Transaction.Type.EXPENSE if amount < 0 else Transaction.Type.INCOME
            category = _default_category(type_)
            if category is None:
                raise ValidationError(
                    "No default category is configured for imported transactions."
                )

            imported_transaction = Transaction.objects.create(
                user=batch.user,
                account=batch.account,
                category=category,
                type=type_,
                amount=abs(amount),
                currency=batch.account.currency,
                date=row_date,
                description=description,
            )
            row.status = ImportRow.Status.IMPORTED
            row.transaction = imported_transaction
            row.save(update_fields=["status", "transaction"])
            imported.append(imported_transaction)

        batch.status = ImportBatch.Status.CONFIRMED
        batch.save(update_fields=["status"])

    return imported


# --- Budget .xlsx import (Section 15/16, budgets variant) -----------------
#
# Fixed header row rather than a column-mapping step: Category, Month,
# Amount, and an optional Household. Re-importing the same
# category+month upserts the existing budget instead of erroring —
# budgets are naturally idempotent per (category, month), unlike
# transactions, so there's no duplicate-detection/skip flow here.

MAX_BUDGET_ROWS = 500
MAX_BUDGET_FILE_SIZE_BYTES = 2 * 1024 * 1024  # 2MB
BUDGET_REQUIRED_HEADERS = ("category", "month", "amount")
BUDGET_OPTIONAL_HEADERS = ("household",)
MONTH_FORMATS = ["%Y-%m-%d", "%Y-%m", "%m/%Y", "%B %Y", "%b %Y"]


def _cell_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _parse_budget_month(raw_value):
    if isinstance(raw_value, datetime):
        return raw_value.date().replace(day=1)
    if isinstance(raw_value, date):
        return raw_value.replace(day=1)

    raw_value = (raw_value or "").strip() if isinstance(raw_value, str) else str(raw_value or "")
    for fmt in MONTH_FORMATS:
        try:
            return datetime.strptime(raw_value, fmt).date().replace(day=1)
        except ValueError:
            continue
    return None


def _parse_budget_amount(raw_value):
    if isinstance(raw_value, (int, float, Decimal)):
        amount = Decimal(str(raw_value))
    else:
        raw_value = (raw_value or "").strip()
        cleaned = raw_value.replace(",", "").replace("₱", "").replace("$", "").strip()
        try:
            amount = Decimal(cleaned)
        except InvalidOperation:
            return None
    if amount <= 0:
        return None
    return amount


def _resolve_budget_category(user, raw_value):
    name = (raw_value or "").strip()
    if not name:
        return None
    from django.db.models import Q

    return (
        Category.objects.filter(Q(is_system=True) | Q(user=user), name__iexact=name)
        .order_by("kind")
        .first()
    )


def _resolve_budget_household(user, raw_value):
    name = (raw_value or "").strip() if raw_value else ""
    if not name:
        return None, True  # (household, is_valid) — blank means "personal", always valid
    household = Household.objects.filter(memberships__user=user, name__iexact=name).first()
    return household, household is not None


def _existing_budget(user, household, category, month):
    qs = Budget.objects.filter(category=category, month=month)
    if household is not None:
        return qs.filter(household=household).first()
    return qs.filter(user=user, household__isnull=True).first()


def create_budget_import_batch(user, uploaded_file):
    if not uploaded_file.name.lower().endswith(".xlsx"):
        raise ValidationError({"file": "Only .xlsx files are accepted."})
    if uploaded_file.size > MAX_BUDGET_FILE_SIZE_BYTES:
        raise ValidationError(
            {"file": f"File exceeds the {MAX_BUDGET_FILE_SIZE_BYTES // 1024}KB limit."}
        )

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()), read_only=True, data_only=True)
        sheet = workbook.active
        all_rows = list(sheet.iter_rows(values_only=True))
    except Exception:
        raise ValidationError({"file": "Could not read this file as a valid .xlsx workbook."})

    if not all_rows:
        raise ValidationError({"file": "The file has no header row."})

    header = [(_cell_text(cell)).lower() for cell in all_rows[0]]
    column_index = {name: i for i, name in enumerate(header) if name}
    missing = [h for h in BUDGET_REQUIRED_HEADERS if h not in column_index]
    if missing:
        raise ValidationError(
            {"file": f"Missing required column(s): {', '.join(missing)}. Expected a header row "
                     f"with Category, Month, Amount (and optionally Household)."}
        )

    data_rows = [row for row in all_rows[1:] if any(cell is not None for cell in row)]
    if not data_rows:
        raise ValidationError({"file": "The file has no data rows."})
    if len(data_rows) > MAX_BUDGET_ROWS:
        raise ValidationError(
            {"file": f"This file has {len(data_rows)} rows; the limit for a single import is {MAX_BUDGET_ROWS}."}
        )

    with db_transaction.atomic():
        batch = BudgetImportBatch.objects.create(
            user=user, filename=uploaded_file.name, row_count=len(data_rows)
        )
        for offset, row in enumerate(data_rows):
            row_number = offset + 2  # header is row 1
            raw_data = {
                "category": _cell_text(row[column_index["category"]]) if column_index["category"] < len(row) else "",
                "month": _cell_text(row[column_index["month"]]) if column_index["month"] < len(row) else "",
                "amount": _cell_text(row[column_index["amount"]]) if column_index["amount"] < len(row) else "",
                "household": (
                    _cell_text(row[column_index["household"]])
                    if "household" in column_index and column_index["household"] < len(row)
                    else ""
                ),
            }
            _stage_budget_row(batch, user, row_number, raw_data)

    return batch


def _stage_budget_row(batch, user, row_number, raw_data):
    category = _resolve_budget_category(user, raw_data["category"])
    if category is None:
        BudgetImportRow.objects.create(
            batch=batch, row_number=row_number, raw_data=raw_data, status=BudgetImportRow.Status.FAILED,
            error=f"Category '{raw_data['category']}' not found.",
        )
        return

    month = _parse_budget_month(raw_data["month"])
    if month is None:
        BudgetImportRow.objects.create(
            batch=batch, row_number=row_number, raw_data=raw_data, status=BudgetImportRow.Status.FAILED,
            error=f"Couldn't parse '{raw_data['month']}' as a month.",
        )
        return

    amount = _parse_budget_amount(raw_data["amount"])
    if amount is None:
        BudgetImportRow.objects.create(
            batch=batch, row_number=row_number, raw_data=raw_data, status=BudgetImportRow.Status.FAILED,
            error=f"Couldn't parse '{raw_data['amount']}' as a nonzero amount.",
        )
        return

    household, household_valid = _resolve_budget_household(user, raw_data["household"])
    if not household_valid:
        BudgetImportRow.objects.create(
            batch=batch, row_number=row_number, raw_data=raw_data, status=BudgetImportRow.Status.FAILED,
            error=f"Household '{raw_data['household']}' not found, or you're not a member.",
        )
        return

    existing = _existing_budget(user, household, category, month)
    BudgetImportRow.objects.create(
        batch=batch,
        row_number=row_number,
        raw_data=raw_data,
        status=BudgetImportRow.Status.PENDING,
        action=BudgetImportRow.Action.UPDATE if existing else BudgetImportRow.Action.CREATE,
    )


def confirm_budget_import_batch(batch, row_ids=None):
    if batch.status == BudgetImportBatch.Status.CONFIRMED:
        raise ValidationError("This import has already been confirmed.")

    pending_rows = list(batch.rows.filter(status=BudgetImportRow.Status.PENDING))
    pending_ids = {str(row.id) for row in pending_rows}

    if row_ids is None:
        selected_ids = pending_ids
    else:
        selected_ids = {str(i) for i in row_ids}
        unknown = selected_ids - pending_ids
        if unknown:
            raise ValidationError(
                {"row_ids": f"Unknown or non-pending row id(s): {sorted(unknown)}"}
            )

    imported = []
    for row in pending_rows:
        if str(row.id) not in selected_ids:
            row.status = BudgetImportRow.Status.SKIPPED
            row.save(update_fields=["status"])
            continue

        category = _resolve_budget_category(batch.user, row.raw_data["category"])
        month = _parse_budget_month(row.raw_data["month"])
        amount = _parse_budget_amount(row.raw_data["amount"])
        household, _ = _resolve_budget_household(batch.user, row.raw_data["household"])

        try:
            with db_transaction.atomic():
                existing = _existing_budget(batch.user, household, category, month)
                if existing is not None:
                    budget = budget_services.update_budget(
                        existing, actor=batch.user, household=household, category=category,
                        month=month, amount=amount,
                    )
                else:
                    budget = budget_services.create_budget(
                        user=batch.user, household=household, category=category, month=month,
                        amount=amount,
                    )
                row.status = BudgetImportRow.Status.IMPORTED
                row.budget = budget
                row.save(update_fields=["status", "budget"])
                imported.append(budget)
        except (ValidationError, PermissionDenied) as exc:
            row.status = BudgetImportRow.Status.FAILED
            row.error = str(exc.detail if hasattr(exc, "detail") else exc)
            row.save(update_fields=["status", "error"])

    batch.status = BudgetImportBatch.Status.CONFIRMED
    batch.save(update_fields=["status"])

    return imported


def build_budget_import_template():
    """An .xlsx with the expected header row + one example row, as bytes."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Budgets"
    sheet.append(["Category", "Month", "Amount", "Household"])
    sheet.append(["Food", "2026-08", 15000, ""])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
